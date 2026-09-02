"""
Monthly LME zinc price (source: World Bank "Pink Sheet" commodity data) -> Neon Postgres.

ZINC_MCX (scrape_zinc.py) is daily MCX futures in INR/kg, but Upstox only exposes
~1 month of history and every other free source (MCX itself, investing.com, stooq)
blocks scraping -- so there's no free way to backfill years of MCX history. This
script tracks a second, separate material (ZINC_LME: monthly, USD/tonne, London
Metal Exchange spot) which the World Bank publishes going back to 1960, as a
reliable stand-in for historical trend data. See migration
017_zinc_lme_material.sql and raw-material-prices/README.md for the rationale.

The Pink Sheet workbook lives at a URL whose hash changes each update, so this
script first discovers the current link from the World Bank's commodity-markets
page rather than hardcoding it.

Usage:
    export DATABASE_URL="postgresql://user:pass@host/dbname?sslmode=require"
    python scrape_zinc_lme.py                # just the last 2 months (routine run)
    python scrape_zinc_lme.py --since 2023-01 # backfill every month from Jan 2023

Requires: requests, beautifulsoup4, openpyxl, psycopg2-binary
    pip install requests beautifulsoup4 openpyxl psycopg2-binary
"""

import argparse
import io
import sys
from datetime import date

import openpyxl
import requests
from bs4 import BeautifulSoup

COMMODITY_MARKETS_PAGE = "https://www.worldbank.org/en/research/commodity-markets"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    )
}

MATERIAL_CODE = "ZINC_LME"


def find_workbook_url():
    """The Pink Sheet's download link changes (hashed URL) each time World Bank
    republishes it, so scrape the current one off the commodity-markets page."""
    resp = requests.get(COMMODITY_MARKETS_PAGE, headers=HEADERS, timeout=20)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    for a in soup.find_all("a", href=True):
        href = a["href"]
        if href.lower().endswith(".xlsx") and "cmo-historical-data-monthly" in href.lower():
            return href

    sys.exit(
        "Could not find the Pink Sheet monthly-data .xlsx link on the World Bank "
        "commodity-markets page -- the page layout may have changed."
    )


def fetch_zinc_series(since):
    """Returns a list of {price_date, price} dicts, one per month >= `since`."""
    url = find_workbook_url()
    resp = requests.get(url, headers=HEADERS, timeout=60)
    resp.raise_for_status()

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb["Monthly Prices"]

    header = [c.value for c in ws[5]]
    if "Zinc" not in header:
        sys.exit("Could not find a 'Zinc' column in the Pink Sheet -- format may have changed.")
    zinc_col = header.index("Zinc")

    rows = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        period = row[0]  # e.g. '2023M01'
        if not period or "M" not in str(period):
            continue
        year, month = str(period).split("M")
        price_date = date(int(year), int(month), 1)
        if price_date < since:
            continue

        value = row[zinc_col]
        if not isinstance(value, (int, float)):
            continue  # '…' placeholder for months not yet published

        rows.append({"price_date": price_date, "price": float(value)})

    return rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--since",
        help=(
            "Earliest year-month to fetch, as YYYY-MM. Default: the last 2 months "
            "(routine run -- just picks up this/last month's published value). "
            "Pass e.g. --since 2023-01 to backfill further history."
        ),
    )
    args = parser.parse_args()
    if args.since:
        since_year, since_month = args.since.split("-")
        since = date(int(since_year), int(since_month), 1)
    else:
        today = date.today()
        since_month_num = today.month - 1 or 12
        since_year_num = today.year if today.month > 1 else today.year - 1
        since = date(since_year_num, since_month_num, 1)

    from db import upsert_rows

    rows = fetch_zinc_series(since)
    if not rows:
        sys.exit(f"No Zinc data found from {since} onward -- check the workbook format.")

    for r in sorted(rows, key=lambda r: r["price_date"]):
        print(r)

    generic_rows = [
        {
            "price_date": r["price_date"],
            "price": r["price"],
            "source": "worldbank",
            "source_url": COMMODITY_MARKETS_PAGE,
            "metadata": {"unit": "USD/tonne", "exchange": "LME"},
        }
        for r in rows
    ]
    upsert_rows(MATERIAL_CODE, generic_rows)


if __name__ == "__main__":
    main()
